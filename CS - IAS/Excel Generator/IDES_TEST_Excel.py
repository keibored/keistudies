"""
Avalanche Effect Excel Generator
=================================
Generates a fully-formulated Excel workbook with 4 sheets:
  1. DES  – Plaintext Avalanche   (flip each of 64 PT bits, fixed key)
  2. DES  – Key Sensitivity        (flip each of 64 key bits, fixed PT)
  3. IDES – Plaintext Avalanche   (flip each of 64 PT bits, fixed 128-bit key)
  4. IDES – Key Sensitivity        (flip each of 128 key bits, fixed PT)

Usage:
    python avalanche_generator.py

    Edit the INPUT SECTION below to change the plaintext / keys.

Requirements:
    pip install openpyxl

Output:
    avalanche_output.xlsx  (in the same directory)
"""

# ═══════════════════════════════════════════════════════════════════════════════
# INPUT SECTION – change these values as needed
# ═══════════════════════════════════════════════════════════════════════════════

PLAINTEXT   = "123456ABCD132536"          # 16 hex chars  = 64-bit block
DES_KEY     = "AABB09182736CCDD"          # 16 hex chars  = 64-bit DES key
IDES_KEY    = "AABB09182736CCDD1683F84F8C1AD287"  # 32 hex chars = 128-bit IDES key

OUTPUT_FILE = "avalanche_output.xlsx"

# ═══════════════════════════════════════════════════════════════════════════════
# DES IMPLEMENTATION  (standard 64-bit key, 16 rounds, Feistel)
# ═══════════════════════════════════════════════════════════════════════════════

_HEX2BIN = {
    '0':"0000",'1':"0001",'2':"0010",'3':"0011",
    '4':"0100",'5':"0101",'6':"0110",'7':"0111",
    '8':"1000",'9':"1001",'A':"1010",'B':"1011",
    'C':"1100",'D':"1101",'E':"1110",'F':"1111",
}
_BIN2HEX = {v: k for k, v in _HEX2BIN.items()}

def hex2bin(s):
    return "".join(_HEX2BIN[c] for c in s.upper())

def bin2hex(s):
    return "".join(_BIN2HEX[s[i:i+4]] for i in range(0, len(s), 4))

def bin2dec(b):
    d, i = 0, 0
    while b:
        d += (b % 10) * (2 ** i)
        b //= 10
        i += 1
    return d

def dec2bin(n):
    r = bin(n).replace("0b", "") or "0"
    pad = (-len(r)) % 4
    return "0" * pad + r

def _permute(bits, table):
    return "".join(bits[t - 1] for t in table)

def _xor(a, b):
    return "".join("0" if x == y else "1" for x, y in zip(a, b))

def _shift_left(k, n):
    return k[n:] + k[:n]

# ── DES tables ──────────────────────────────────────────────────────────────
_IP  = [58,50,42,34,26,18,10,2, 60,52,44,36,28,20,12,4,
        62,54,46,38,30,22,14,6, 64,56,48,40,32,24,16,8,
        57,49,41,33,25,17, 9,1, 59,51,43,35,27,19,11,3,
        61,53,45,37,29,21,13,5, 63,55,47,39,31,23,15,7]

_FP  = [40,8,48,16,56,24,64,32, 39,7,47,15,55,23,63,31,
        38,6,46,14,54,22,62,30, 37,5,45,13,53,21,61,29,
        36,4,44,12,52,20,60,28, 35,3,43,11,51,19,59,27,
        34,2,42,10,50,18,58,26, 33,1,41, 9,49,17,57,25]

_EXP = [32,1,2,3,4,5, 4,5,6,7,8,9, 8,9,10,11,12,13,
        12,13,14,15,16,17, 16,17,18,19,20,21, 20,21,22,23,24,25,
        24,25,26,27,28,29, 28,29,30,31,32,1]

_P   = [16,7,20,21,29,12,28,17, 1,15,23,26,5,18,31,10,
         2,8,24,14,32,27, 3,9, 19,13,30,6,22,11,4,25]

_PC1 = [57,49,41,33,25,17,9,1,58,50,42,34,26,18,
        10,2,59,51,43,35,27,19,11,3,60,52,44,36,
        63,55,47,39,31,23,15,7,62,54,46,38,30,22,
        14,6,61,53,45,37,29,21,13,5,28,20,12,4]

_PC2 = [14,17,11,24,1,5,3,28,15,6,21,10,23,19,12,4,
        26,8,16,7,27,20,13,2,41,52,31,37,47,55,
        30,40,51,45,33,48,44,49,39,56,34,53,
        46,42,50,36,29,32]

_SHIFTS = [1,1,2,2,2,2,2,2,1,2,2,2,2,2,2,1]

_SBOX = [
    [[14,4,13,1,2,15,11,8,3,10,6,12,5,9,0,7],
     [0,15,7,4,14,2,13,1,10,6,12,11,9,5,3,8],
     [4,1,14,8,13,6,2,11,15,12,9,7,3,10,5,0],
     [15,12,8,2,4,9,1,7,5,11,3,14,10,0,6,13]],
    [[15,1,8,14,6,11,3,4,9,7,2,13,12,0,5,10],
     [3,13,4,7,15,2,8,14,12,0,1,10,6,9,11,5],
     [0,14,7,11,10,4,13,1,5,8,12,6,9,3,2,15],
     [13,8,10,1,3,15,4,2,11,6,7,12,0,5,14,9]],
    [[10,0,9,14,6,3,15,5,1,13,12,7,11,4,2,8],
     [13,7,0,9,3,4,6,10,2,8,5,14,12,11,15,1],
     [13,6,4,9,8,15,3,0,11,1,2,12,5,10,14,7],
     [1,10,13,0,6,9,8,7,4,15,14,3,11,5,2,12]],
    [[7,13,14,3,0,6,9,10,1,2,8,5,11,12,4,15],
     [13,8,11,5,6,15,0,3,4,7,2,12,1,10,14,9],
     [10,6,9,0,12,11,7,13,15,1,3,14,5,2,8,4],
     [3,15,0,6,10,1,13,8,9,4,5,11,12,7,2,14]],
    [[2,12,4,1,7,10,11,6,8,5,3,15,13,0,14,9],
     [14,11,2,12,4,7,13,1,5,0,15,10,3,9,8,6],
     [4,2,1,11,10,13,7,8,15,9,12,5,6,3,0,14],
     [11,8,12,7,1,14,2,13,6,15,0,9,10,4,5,3]],
    [[12,1,10,15,9,2,6,8,0,13,3,4,14,7,5,11],
     [10,15,4,2,7,12,9,5,6,1,13,14,0,11,3,8],
     [9,14,15,5,2,8,12,3,7,0,4,10,1,13,11,6],
     [4,3,2,12,9,5,15,10,11,14,1,7,6,0,8,13]],
    [[4,11,2,14,15,0,8,13,3,12,9,7,5,10,6,1],
     [13,0,11,7,4,9,1,10,14,3,5,12,2,15,8,6],
     [1,4,11,13,12,3,7,14,10,15,6,8,0,5,9,2],
     [6,11,13,8,1,4,10,7,9,5,0,15,14,2,3,12]],
    [[13,2,8,4,6,15,11,1,10,9,3,14,5,0,12,7],
     [1,15,13,8,10,3,7,4,12,5,6,11,0,14,9,2],
     [7,11,4,1,9,12,14,2,0,6,10,13,15,3,5,8],
     [2,1,14,7,4,10,8,13,15,12,9,0,3,5,6,11]],
]

def _des_round_keys(key_hex):
    """Generate 16 DES round keys from a 64-bit hex key string."""
    k = _permute(hex2bin(key_hex), _PC1)
    L, R = k[:28], k[28:]
    round_keys = []
    for s in _SHIFTS:
        L = _shift_left(L, s)
        R = _shift_left(R, s)
        round_keys.append(_permute(L + R, _PC2))
    return round_keys

def _des_encrypt_bin(pt_bin, round_keys):
    """Encrypt a 64-bit binary string using provided round keys."""
    pt = _permute(pt_bin, _IP)
    L, R = pt[:32], pt[32:]
    for rk in round_keys:
        exp_R   = _permute(R, _EXP)
        xored   = _xor(exp_R, rk)
        sb = ""
        for j in range(8):
            chunk = xored[j*6:(j+1)*6]
            row = bin2dec(int(chunk[0] + chunk[5]))
            col = bin2dec(int(chunk[1:5]))
            sb += dec2bin(_SBOX[j][row][col])
        sb = _permute(sb, _P)
        L, R = R, _xor(L, sb)
    return bin2hex(_permute(R + L, _FP))

def des_encrypt(pt_hex, key_hex):
    rk = _des_round_keys(key_hex)
    return _des_encrypt_bin(hex2bin(pt_hex), rk)

def des_encrypt_with_key_rks(pt_hex, round_keys):
    return _des_encrypt_bin(hex2bin(pt_hex), round_keys)

# ═══════════════════════════════════════════════════════════════════════════════
# IDES IMPLEMENTATION  — delegates to improved_des.py
# (SHA-256 key schedule, dynamic S-boxes, 32 rounds)
# ═══════════════════════════════════════════════════════════════════════════════

from improved_des import encrypt as _ides_encrypt_real

def ides_encrypt(pt_hex, key128_hex):
    """Wrapper: calls the real IDES (improved_des.py) encrypt function."""
    return _ides_encrypt_real(pt_hex.upper(), key128_hex.upper())

# ═══════════════════════════════════════════════════════════════════════════════
# BIT-FLIP HELPER
# ═══════════════════════════════════════════════════════════════════════════════

def flip_bit(hex_str, bit_pos):
    """Flip bit at 1-indexed position in hex string."""
    bits = list(hex2bin(hex_str))
    bits[bit_pos - 1] = '1' if bits[bit_pos - 1] == '0' else '0'
    return bin2hex("".join(bits))

# ═══════════════════════════════════════════════════════════════════════════════
# DATA GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def generate_pt_avalanche_data(pt_hex, encrypt_fn):
    """
    Flip each of 64 plaintext bits. Returns list of:
      (bit_num, orig_pt, mod_pt, orig_ct, mod_ct)
    """
    orig_ct = encrypt_fn(pt_hex)
    rows = []
    for bit in range(1, 65):
        mod_pt = flip_bit(pt_hex, bit)
        mod_ct = encrypt_fn(mod_pt)
        rows.append((bit, pt_hex, mod_pt, orig_ct, mod_ct))
    return rows

def generate_key_avalanche_data(pt_hex, key_hex, key_bits, encrypt_fn):
    """
    Flip each key bit. Returns list of:
      (bit_num, orig_key, mod_key, orig_ct, mod_ct)
    key_bits = 64 for DES, 128 for IDES
    """
    orig_ct = encrypt_fn(pt_hex)
    rows = []
    for bit in range(1, key_bits + 1):
        mod_key = flip_bit(key_hex, bit)
        mod_ct  = encrypt_fn(pt_hex, mod_key=mod_key)
        rows.append((bit, key_hex, mod_key, orig_ct, mod_ct))
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL SHEET BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ── Column layout for a PT-avalanche sheet (orig string = 16 hex, key = any) ──
#
#  A   : Original PT/Key string  (col 1)
#  B   : Flipped bit #           (col 2)
#  C…  : Orig Hex  (hex_n cols)  starting col 3,   gap after
#  …   : Orig 4-bit (hex_n cols) starting col 3+hex_n+1 (gap=1)
#  …   : Orig Bits  (hex_n*4)    starting col 3+hex_n*2+2 (gap=1)
#  CW  : Mod string col           starting col 3+hex_n*5+3  (gap=1)
#  …   : Mod  Hex  (hex_n)
#  …   : Mod  4-bit (hex_n)       gap=1
#  …   : Mod  Bits  (hex_n*4)     gap=1
#  GR  : Orig CT string col       gap=1
#  …   : Orig CT Hex (16)
#  …   : Orig CT 4-bit (16)       gap=1
#  …   : Orig CT Bits (64)        gap=1
#  KM  : Mod CT string col        gap=1
#  …   : Mod CT Hex (16)
#  …   : Mod CT 4-bit (16)        gap=1
#  …   : Mod CT Bits (64)         gap=1
#  …   : XOR (64)                 gap=1
#  …   : Hamming (1)              gap=1
#
# For PT avalanche: hex_n = 16
# For Key sensitivity (DES):  hex_n_key=16, same layout as PT avalanche
# For Key sensitivity (IDES): hex_n_key=32, layout scales accordingly

def _col_layout(hex_n_orig):
    """
    Compute column start positions given number of orig hex chars.
    CT is always 16 hex chars (64-bit block output).
    """
    GAP = 1
    c = {}
    c['A']               = 1
    c['B']               = 2
    c['ORIG_HEX']        = 3                                           # hex_n cols
    c['ORIG_4BIT']       = c['ORIG_HEX']   + hex_n_orig + GAP        # hex_n cols
    c['ORIG_BITS']       = c['ORIG_4BIT']  + hex_n_orig + GAP        # hex_n*4 cols
    c['MOD_STR']         = c['ORIG_BITS']  + hex_n_orig*4 + GAP      # 1 col (string)
    c['MOD_HEX']         = c['MOD_STR']   + 1                        # hex_n cols
    c['MOD_4BIT']        = c['MOD_HEX']   + hex_n_orig + GAP        # hex_n cols
    c['MOD_BITS']        = c['MOD_4BIT']  + hex_n_orig + GAP        # hex_n*4 cols
    c['ORIGCT_STR']      = c['MOD_BITS']  + hex_n_orig*4 + GAP      # 1 col (string)
    c['ORIGCT_HEX']      = c['ORIGCT_STR'] + 1                       # 16 cols
    c['ORIGCT_4BIT']     = c['ORIGCT_HEX'] + 16 + GAP               # 16 cols
    c['ORIGCT_BITS']     = c['ORIGCT_4BIT'] + 16 + GAP              # 64 cols
    c['MODCT_STR']       = c['ORIGCT_BITS'] + 64 + GAP              # 1 col (string)
    c['MODCT_HEX']       = c['MODCT_STR']  + 1                       # 16 cols
    c['MODCT_4BIT']      = c['MODCT_HEX']  + 16 + GAP               # 16 cols
    c['MODCT_BITS']      = c['MODCT_4BIT'] + 16 + GAP               # 64 cols
    c['XOR']             = c['MODCT_BITS'] + 64 + GAP               # 64 cols
    c['HAMMING']         = c['XOR']        + 64 + GAP               # 1 col
    return c

def _cl(col_idx):
    return get_column_letter(col_idx)

def build_sheet(wb, sheet_name,
                rows,           # list of (bit, orig_str, mod_str, orig_ct, mod_ct)
                hex_n,          # number of hex chars in orig string (16 or 32)
                label_a,        # "Original Plaintext" / "Original Key"
                label_b,        # "Flipped PT Bit #" / "Flipped Key Bit #"
                label_orig_hex, # "Original Plaintext (Hex)" etc.
                label_orig_4bit,
                label_orig_bits,
                label_mod_str,
                label_mod_hex,
                label_mod_4bit,
                label_mod_bits,
                ):
    ws = wb.create_sheet(sheet_name)
    C  = _col_layout(hex_n)

    # ── HEADER ROW 1 ──────────────────────────────────────────────────────
    headers1 = {
        C['A']:           label_a,
        C['B']:           label_b,
        C['ORIG_HEX']:    label_orig_hex,
        C['ORIG_4BIT']:   label_orig_4bit,
        C['ORIG_BITS']:   label_orig_bits,
        C['MOD_STR']:     label_mod_str,
        C['MOD_HEX']:     label_mod_hex,
        C['MOD_4BIT']:    label_mod_4bit,
        C['MOD_BITS']:    label_mod_bits,
        C['ORIGCT_STR']:  'Original Ciphertext',
        C['ORIGCT_HEX']:  'Original Ciphertext (Hex)',
        C['ORIGCT_4BIT']: 'Original CT (4-bit)',
        C['ORIGCT_BITS']: 'Original CT (Bits)',
        C['MODCT_STR']:   'Modified Ciphertext',
        C['MODCT_HEX']:   'Modified Ciphertext (Hex)',
        C['MODCT_4BIT']:  'Modified CT (4-bit)',
        C['MODCT_BITS']:  'Modified CT (Bits)',
        C['XOR']:         'XOR (Orig CT \u2295 Mod CT bits)',
        C['HAMMING']:     'Hamming Distance',
    }
    for col, val in headers1.items():
        ws.cell(row=1, column=col, value=val)

    # ── HEADER ROW 2 ──────────────────────────────────────────────────────
    for i in range(hex_n):
        ws.cell(row=2, column=C['ORIG_4BIT'] + i, value=f'N{i+1}')
    for i in range(hex_n * 4):
        ws.cell(row=2, column=C['ORIG_BITS'] + i, value=f'b{i+1}')
    for i in range(hex_n):
        ws.cell(row=2, column=C['MOD_4BIT'] + i, value=f'N{i+1}')
    for i in range(hex_n * 4):
        ws.cell(row=2, column=C['MOD_BITS'] + i, value=f'b{i+1}')
    for i in range(16):
        ws.cell(row=2, column=C['ORIGCT_4BIT'] + i, value=f'N{i+1}')
    for i in range(64):
        ws.cell(row=2, column=C['ORIGCT_BITS'] + i, value=f'b{i+1}')
    for i in range(16):
        ws.cell(row=2, column=C['MODCT_4BIT'] + i, value=f'N{i+1}')
    for i in range(64):
        ws.cell(row=2, column=C['MODCT_BITS'] + i, value=f'b{i+1}')
    for i in range(64):
        ws.cell(row=2, column=C['XOR'] + i, value=f'b{i+1}')

    # ── DATA ROWS ─────────────────────────────────────────────────────────
    for idx, (bit_num, orig_str, mod_str, orig_ct, mod_ct) in enumerate(rows):
        r = idx + 3

        # Col A: original string, Col B: flipped bit number
        ws.cell(row=r, column=C['A'], value=orig_str)
        ws.cell(row=r, column=C['B'], value=bit_num)

        # ── Orig Hex: MID($A, n, 1) ──
        for i in range(hex_n):
            ws.cell(row=r, column=C['ORIG_HEX'] + i,
                    value=f'=MID($A{r},{i+1},1)')

        # ── Orig 4-bit: DEC2BIN(HEX2DEC(hex_cell), 4) ──
        for i in range(hex_n):
            hc = _cl(C['ORIG_HEX'] + i)
            ws.cell(row=r, column=C['ORIG_4BIT'] + i,
                    value=f'=DEC2BIN(HEX2DEC({hc}{r}),4)')

        # ── Orig Bits: MID(4bit_cell, j, 1) ──
        for i in range(hex_n):
            fc = _cl(C['ORIG_4BIT'] + i)
            for j in range(4):
                ws.cell(row=r, column=C['ORIG_BITS'] + i*4 + j,
                        value=f'=MID({fc}{r},{j+1},1)')

        # ── Mod string (literal value) ──
        ws.cell(row=r, column=C['MOD_STR'], value=mod_str)

        # ── Mod Hex: MID($MOD_STR_col, n, 1) ──
        msc = _cl(C['MOD_STR'])
        for i in range(hex_n):
            ws.cell(row=r, column=C['MOD_HEX'] + i,
                    value=f'=MID(${msc}{r},{i+1},1)')

        # ── Mod 4-bit ──
        for i in range(hex_n):
            hc = _cl(C['MOD_HEX'] + i)
            ws.cell(row=r, column=C['MOD_4BIT'] + i,
                    value=f'=DEC2BIN(HEX2DEC({hc}{r}),4)')

        # ── Mod Bits ──
        for i in range(hex_n):
            fc = _cl(C['MOD_4BIT'] + i)
            for j in range(4):
                ws.cell(row=r, column=C['MOD_BITS'] + i*4 + j,
                        value=f'=MID({fc}{r},{j+1},1)')

        # ── Orig CT string (literal) ──
        ws.cell(row=r, column=C['ORIGCT_STR'], value=orig_ct)

        # ── Orig CT Hex: MID($ORIGCT_STR_col, n, 1) ──
        osc = _cl(C['ORIGCT_STR'])
        for i in range(16):
            ws.cell(row=r, column=C['ORIGCT_HEX'] + i,
                    value=f'=MID(${osc}{r},{i+1},1)')

        # ── Orig CT 4-bit ──
        for i in range(16):
            hc = _cl(C['ORIGCT_HEX'] + i)
            ws.cell(row=r, column=C['ORIGCT_4BIT'] + i,
                    value=f'=DEC2BIN(HEX2DEC({hc}{r}),4)')

        # ── Orig CT Bits ──
        for i in range(16):
            fc = _cl(C['ORIGCT_4BIT'] + i)
            for j in range(4):
                ws.cell(row=r, column=C['ORIGCT_BITS'] + i*4 + j,
                        value=f'=MID({fc}{r},{j+1},1)')

        # ── Mod CT string (literal) ──
        ws.cell(row=r, column=C['MODCT_STR'], value=mod_ct)

        # ── Mod CT Hex: MID($MODCT_STR_col, n, 1) ──
        mctsc = _cl(C['MODCT_STR'])
        for i in range(16):
            ws.cell(row=r, column=C['MODCT_HEX'] + i,
                    value=f'=MID(${mctsc}{r},{i+1},1)')

        # ── Mod CT 4-bit ──
        for i in range(16):
            hc = _cl(C['MODCT_HEX'] + i)
            ws.cell(row=r, column=C['MODCT_4BIT'] + i,
                    value=f'=DEC2BIN(HEX2DEC({hc}{r}),4)')

        # ── Mod CT Bits ──
        for i in range(16):
            fc = _cl(C['MODCT_4BIT'] + i)
            for j in range(4):
                ws.cell(row=r, column=C['MODCT_BITS'] + i*4 + j,
                        value=f'=MID({fc}{r},{j+1},1)')

        # ── XOR: MOD(origct_bit + modct_bit, 2) ──
        for i in range(64):
            oc = _cl(C['ORIGCT_BITS'] + i)
            mc = _cl(C['MODCT_BITS']  + i)
            ws.cell(row=r, column=C['XOR'] + i,
                    value=f'=MOD({oc}{r}+{mc}{r},2)')

        # ── Hamming Distance: SUM(XOR range) ──
        x0 = _cl(C['XOR'])
        x1 = _cl(C['XOR'] + 63)
        ws.cell(row=r, column=C['HAMMING'],
                value=f'=SUM({x0}{r}:{x1}{r})')

    print(f"  Built: '{sheet_name}'  ({len(rows)} data rows, "
          f"{C['HAMMING']} total columns)")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def build_summary_sheet(wb):
    """
    Add a Summary sheet with VLOOKUP formulas pulling Hamming distances
    from all 4 data sheets, displayed side by side with their means.

    Sheet layout (no colors, no styles):
      Row 1: headers
      Row 2 onwards: one row per bit flip

    Columns:
      A  – Bit #  (1..128; DES sheets only go to 64)
      B  – DES PT Avalanche    Hamming  (VLOOKUP from 'DES – Plaintext Avalanche')
      C  – DES Key Sensitivity Hamming  (VLOOKUP from 'DES – Key Sensitivity')
      D  – IDES PT Avalanche   Hamming  (VLOOKUP from 'IDES – Plaintext Avalanche')
      E  – IDES Key Sensitivity Hamming (VLOOKUP from 'IDES – Key Sensitivity')
    Row after last data row:
      A  – "Mean"
      B..E – AVERAGE of each column
    """
    ws = wb.create_sheet('Summary')

    # ── Column indices computed from _col_layout ──────────────────────────
    # hex_n=16  → HAMMING col 467 (QY); table starts at col B (2) → index = 466
    # hex_n=32  → HAMMING col 659 (YI); table starts at col B (2) → index = 658
    DES_VLOOKUP_IDX  = 466   # for sheets with hex_n=16
    IDES_KEY_IDX     = 658   # for 'IDES – Key Sensitivity' (hex_n=32)

    # Sheet names (must match exactly)
    SH_DES_PT   = "'DES \u2013 Plaintext Avalanche'"
    SH_DES_KEY  = "'DES \u2013 Key Sensitivity'"
    SH_IDES_PT  = "'IDES \u2013 Plaintext Avalanche'"
    SH_IDES_KEY = "'IDES \u2013 Key Sensitivity'"

    # ── Header row ────────────────────────────────────────────────────────
    ws['A1'] = 'Bit #'
    ws['B1'] = 'DES PT Avalanche Hamming'
    ws['C1'] = 'DES Key Sensitivity Hamming'
    ws['D1'] = 'IDES PT Avalanche Hamming'
    ws['E1'] = 'IDES Key Sensitivity Hamming'

    # ── Data rows (max 128 bits; DES sheets only have 64 rows → IFERROR) ──
    MAX_BITS = 128
    for bit in range(1, MAX_BITS + 1):
        r = bit + 1   # row 2 = bit 1

        ws.cell(row=r, column=1, value=bit)

        # DES PT Avalanche (hex_n=16, 64 rows, data rows 3..66)
        ws.cell(row=r, column=2,
                value=(f'=IFERROR(VLOOKUP(A{r},'
                       f'{SH_DES_PT}!$B$3:$QY$66,'
                       f'{DES_VLOOKUP_IDX},0),"")'))

        # DES Key Sensitivity (hex_n=16, 64 rows, data rows 3..66)
        ws.cell(row=r, column=3,
                value=(f'=IFERROR(VLOOKUP(A{r},'
                       f'{SH_DES_KEY}!$B$3:$QY$66,'
                       f'{DES_VLOOKUP_IDX},0),"")'))

        # IDES PT Avalanche (hex_n=16, 64 rows, data rows 3..66)
        ws.cell(row=r, column=4,
                value=(f'=IFERROR(VLOOKUP(A{r},'
                       f'{SH_IDES_PT}!$B$3:$QY$66,'
                       f'{DES_VLOOKUP_IDX},0),"")'))

        # IDES Key Sensitivity (hex_n=32, 128 rows, data rows 3..130)
        ws.cell(row=r, column=5,
                value=(f'=VLOOKUP(A{r},{SH_IDES_KEY}!$B$3:$YI$130,'
                       f'{IDES_KEY_IDX},0)'))

    # ── Mean row ──────────────────────────────────────────────────────────
    mean_row = MAX_BITS + 2
    ws.cell(row=mean_row, column=1, value='Mean')
    ws.cell(row=mean_row, column=2, value='=AVERAGE(B2:B65)')   # DES PT: bits 1-64
    ws.cell(row=mean_row, column=3, value='=AVERAGE(C2:C65)')   # DES Key: bits 1-64
    ws.cell(row=mean_row, column=4, value='=AVERAGE(D2:D65)')   # IDES PT: bits 1-64
    ws.cell(row=mean_row, column=5, value='=AVERAGE(E2:E129)')  # IDES Key: bits 1-128

    print("  Built: 'Summary'  (128 bit rows + Mean row, VLOOKUP-based)")
    return ws


def main():
    pt  = PLAINTEXT.upper()
    dk  = DES_KEY.upper()
    ik  = IDES_KEY.upper()

    # Validate inputs
    assert len(pt) == 16,  f"PLAINTEXT must be 16 hex chars, got {len(pt)}"
    assert len(dk) == 16,  f"DES_KEY must be 16 hex chars, got {len(dk)}"
    assert len(ik) == 32,  f"IDES_KEY must be 32 hex chars, got {len(ik)}"
    for s, name in [(pt,'PLAINTEXT'),(dk,'DES_KEY'),(ik,'IDES_KEY')]:
        bad = [c for c in s if c not in _HEX2BIN]
        assert not bad, f"{name} contains invalid hex chars: {bad}"

    print("=" * 60)
    print("Avalanche Effect Excel Generator")
    print("=" * 60)
    print(f"  Plaintext : {pt}")
    print(f"  DES Key   : {dk}")
    print(f"  IDES Key  : {ik}")
    print()

    # ── Encrypt functions ──────────────────────────────────────────────────
    def des_enc(pt_hex, mod_key=None):
        return des_encrypt(pt_hex, mod_key if mod_key else dk)

    def ides_enc(pt_hex, mod_key=None):
        return ides_encrypt(pt_hex, mod_key if mod_key else ik)

    # ── Generate data ──────────────────────────────────────────────────────
    print("Generating DES Plaintext Avalanche data...")
    des_pt_rows  = generate_pt_avalanche_data(pt, des_enc)

    print("Generating DES Key Sensitivity data...")
    des_key_rows = generate_key_avalanche_data(pt, dk, 64, des_enc)

    print("Generating IDES Plaintext Avalanche data...")
    ides_pt_rows  = generate_pt_avalanche_data(pt, ides_enc)

    print("Generating IDES Key Sensitivity data...")
    ides_key_rows = generate_key_avalanche_data(pt, ik, 128, ides_enc)

    # ── Build workbook ─────────────────────────────────────────────────────
    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # 1. DES – Plaintext Avalanche (16 hex, 64 bit flips)
    build_sheet(
        wb, 'DES – Plaintext Avalanche',
        rows            = des_pt_rows,
        hex_n           = 16,
        label_a         = 'Original Plaintext',
        label_b         = 'Flipped PT Bit #',
        label_orig_hex  = 'Original Plaintext (Hex)',
        label_orig_4bit = 'Original Plaintext (4-bit)',
        label_orig_bits = 'Original Plaintext (Bits)',
        label_mod_str   = 'Modified Plaintext',
        label_mod_hex   = 'Modified Plaintext (Hex)',
        label_mod_4bit  = 'Modified Plaintext (4-bit)',
        label_mod_bits  = 'Modified Plaintext (Bits)',
    )

    # 2. DES – Key Sensitivity (16 hex key, 64 bit flips)
    build_sheet(
        wb, 'DES – Key Sensitivity',
        rows            = des_key_rows,
        hex_n           = 16,
        label_a         = 'Original Key',
        label_b         = 'Flipped Key Bit #',
        label_orig_hex  = 'Original Key (Hex)',
        label_orig_4bit = 'Original Key (4-bit)',
        label_orig_bits = 'Original Key (Bits)',
        label_mod_str   = 'Modified Key',
        label_mod_hex   = 'Modified Key (Hex)',
        label_mod_4bit  = 'Modified Key (4-bit)',
        label_mod_bits  = 'Modified Key (Bits)',
    )

    # 3. IDES – Plaintext Avalanche (16 hex PT, 64 bit flips)
    build_sheet(
        wb, 'IDES – Plaintext Avalanche',
        rows            = ides_pt_rows,
        hex_n           = 16,
        label_a         = 'Original Plaintext',
        label_b         = 'Flipped PT Bit #',
        label_orig_hex  = 'Original Plaintext (Hex)',
        label_orig_4bit = 'Original Plaintext (4-bit)',
        label_orig_bits = 'Original Plaintext (Bits)',
        label_mod_str   = 'Modified Plaintext',
        label_mod_hex   = 'Modified Plaintext (Hex)',
        label_mod_4bit  = 'Modified Plaintext (4-bit)',
        label_mod_bits  = 'Modified Plaintext (Bits)',
    )

    # 4. IDES – Key Sensitivity (32 hex key, 128 bit flips)
    build_sheet(
        wb, 'IDES – Key Sensitivity',
        rows            = ides_key_rows,
        hex_n           = 32,
        label_a         = 'Original Key',
        label_b         = 'Flipped Key Bit #',
        label_orig_hex  = 'Original Key (Hex)',
        label_orig_4bit = 'Original Key (4-bit)',
        label_orig_bits = 'Original Key (Bits)',
        label_mod_str   = 'Modified Key',
        label_mod_hex   = 'Modified Key (Hex)',
        label_mod_4bit  = 'Modified Key (4-bit)',
        label_mod_bits  = 'Modified Key (Bits)',
    )

    # 5. Summary – VLOOKUP-based side-by-side Hamming distances + means
    build_summary_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")

    # ── Summary stats ──────────────────────────────────────────────────────
    print("\nAvalanche Effect Summary")
    print("-" * 50)
    for label, rows in [
        ("DES  PT Avalanche", des_pt_rows),
        ("DES  Key Sensitivity", des_key_rows),
        ("IDES PT Avalanche", ides_pt_rows),
        ("IDES Key Sensitivity", ides_key_rows),
    ]:
        hds = [bin(int(r[4], 16) ^ int(r[3], 16)).count('1') for r in rows]
        avg = sum(hds) / len(hds)
        print(f"  {label:25s}  avg Hamming = {avg:.2f}/64  "
              f"({avg/64*100:.1f}%)")
    print()


if __name__ == "__main__":
    main()