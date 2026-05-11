"""
ides_process.py — IDES Complete Step-by-Step Process Visualizer  (FULL REWRITE)

Generates a detailed Excel workbook showing the full bit-by-bit process for
both IDES encryption and decryption, matching the professor's DES reference format.

Improvements over previous version:
  [1] KEY SCHEDULE — 112-bit concat row added (PC-1(KL) || PC-1(KR)) before SHA-256
  [2] KEY SCHEDULE Step 4 — all 32 rounds with per-round substeps:
        rotate C/D → show C[:28], D[:28] selection → CD56 → PC-2 → SK
  [3] S-BOXES — full generation trace: master seed, per-row seed, Fisher-Yates
        shuffle output, NL score, attempt number — then final 4×16 grid
  [4] ROUND SHEETS — reference-format E-expansion table (Block / BorrowedLeft /
        Original-4-bit / BorrowedRight / 6-bit result), full XOR/S-box/P/XOR detail
  [5] ENTIRE WORKBOOK — bit-per-cell layout throughout (no long binary strings in
        merged cells); IP/FP shown as 8×8 grids; S-boxes as 4×16 grids;
        sub-keys as 48 individual bit cells; PC-1/PC-2 as proper tables

Sheets generated:
  IO SHEET     — Overview with 8×8 IP/FP grids, bit-per-cell round states
  KEY SCHEDULE — PC-1 grid, 112-bit concat row, SHA-256, Step-4 per-round detail
  S-BOXES      — Generation trace + 4×16 final tables
  R1 … R32     — Encryption rounds (reference-format E expansion, full detail)
  D1 … D32     — Decryption rounds (same format, reversed sub-keys)
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import hex_to_bits, bits_to_hex, permute, xor_bits, left_rotate, bits_to_bytes, sha256_hash
from key_schedule import build_key_schedule, PC1, PC2, SHIFT_SCHEDULE
from sbox_generator import build_sboxes, fisher_yates_shuffle, _row_nonlinearity
from feistel import f_function, E_TABLE, P_TABLE
from improved_des import IP_TABLE, FP_TABLE, encrypt as ides_encrypt, decrypt as ides_decrypt


# ============================================================
# COLOURS & FONTS
# ============================================================
_DK_BLUE  = PatternFill("solid", fgColor="1F4E79")
_MED_BLUE = PatternFill("solid", fgColor="2E75B6")
_LT_BLUE  = PatternFill("solid", fgColor="DDEBF7")
_YELLOW   = PatternFill("solid", fgColor="FFFF00")
_GREEN    = PatternFill("solid", fgColor="E2EFDA")
_ORANGE   = PatternFill("solid", fgColor="FCE4D6")
_PURPLE   = PatternFill("solid", fgColor="E8D5F5")
_RED_LT   = PatternFill("solid", fgColor="FFE0E0")
_GREY     = PatternFill("solid", fgColor="F2F2F2")
_TEAL     = PatternFill("solid", fgColor="D6EAF8")
_CREAM    = PatternFill("solid", fgColor="FFFACD")

_WH_BOLD  = Font(bold=True, color="FFFFFF", size=9)
_BK_BOLD  = Font(bold=True, color="000000", size=9)
_SMALL    = Font(size=8)
_TINY     = Font(size=7, italic=True)
_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT     = Alignment(horizontal="left",   vertical="center")

thin = Side(style="thin")
_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================
# LOW-LEVEL CELL HELPERS
# ============================================================

def _c(ws, row, col, val="", fill=None, font=None, align=None, border=False):
    cell = ws.cell(row=row, column=col, value=val)
    if fill:   cell.fill  = fill
    if font:   cell.font  = font
    else:      cell.font  = _SMALL
    if align:  cell.alignment = align
    else:      cell.alignment = _CENTER
    if border: cell.border = _BORDER
    return cell


def _hdr(ws, row, col, val, fill=_DK_BLUE, span=1):
    """Dark-blue section header, optionally spanning columns."""
    _c(ws, row, col, val, fill, _WH_BOLD, _CENTER)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + span - 1)


def _sec(ws, row, col, val, span=1, fill=_MED_BLUE):
    _hdr(ws, row, col, val, fill=fill, span=span)


def _col_hdr(ws, row, col, val, fill=_GREY):
    """Column label header cell."""
    _c(ws, row, col, val, fill, _BK_BOLD, _CENTER, border=True)


def _bit_label_row(ws, row, start_col, n, fill=_GREY):
    """Write bit-position labels 1..n in individual cells."""
    for i in range(n):
        c = ws.cell(row=row, column=start_col + i, value=i + 1)
        c.fill = fill
        c.font = _TINY
        c.alignment = _CENTER


def _bits_row(ws, row, start_col, bits, fill=None, border=False):
    """Write each bit in its own cell."""
    for i, b in enumerate(bits):
        _c(ws, row, start_col + i, b, fill, _SMALL, _CENTER, border)


def _narrow_cols(ws, start_col, end_col, width=2.2):
    for c in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _bstr(bits): return ''.join(str(b) for b in bits)
def _bhex(bits): return bits_to_hex(bits)


# ============================================================
# PERMUTATION TABLE GRIDS
# ============================================================

def _write_perm_grid(ws, row, col, table, nrows, ncols, title, fill=_GREY):
    """
    Display a flat permutation table as an (nrows × ncols) grid with
    row-index labels in column A and column-index labels across the top.
    """
    _sec(ws, row, col, title, span=ncols + 1, fill=_MED_BLUE); row += 1
    _c(ws, row, col, "Row \\ Col", _GREY, _BK_BOLD, _CENTER)
    for ci in range(ncols):
        _col_hdr(ws, row, col + 1 + ci, ci + 1)
    row += 1
    for ri in range(nrows):
        _c(ws, row, col, ri + 1, _GREY, _BK_BOLD, _CENTER)
        for ci in range(ncols):
            idx = ri * ncols + ci
            v = table[idx] if idx < len(table) else ""
            _c(ws, row, col + 1 + ci, v, fill, _SMALL, _CENTER, border=True)
        row += 1
    return row + 1


def _write_ip_grid(ws, row, col, table, title):
    """8×8 grid for the 64-entry IP / FP tables."""
    _sec(ws, row, col, title, span=9, fill=_MED_BLUE); row += 1
    _c(ws, row, col, "Row \\ Col", _GREY, _BK_BOLD, _CENTER)
    for ci in range(8):
        _col_hdr(ws, row, col + 1 + ci, ci + 1)
    row += 1
    for ri in range(8):
        _c(ws, row, col, ri + 1, _GREY, _BK_BOLD, _CENTER)
        for ci in range(8):
            v = table[ri * 8 + ci]
            _c(ws, row, col + 1 + ci, v, _LT_BLUE, _SMALL, _CENTER, border=True)
        row += 1
    return row + 1


# ============================================================
# COMPUTE ALL ROUND STATES
# ============================================================

def compute_all_states(pt_bits, key_bits):
    sub_keys = build_key_schedule(key_bits)
    sboxes   = build_sboxes(key_bits)

    after_ip = permute(pt_bits, IP_TABLE)
    L, R = after_ip[:32], after_ip[32:]
    enc_states = [(list(L), list(R))]
    for i in range(32):
        F     = f_function(R, sub_keys[i], sboxes)
        new_R = xor_bits(L, F)
        L, R  = R, new_R
        enc_states.append((list(L), list(R)))

    pre_fp  = enc_states[32][1] + enc_states[32][0]
    ct_bits = permute(pre_fp, FP_TABLE)

    rev_keys = list(reversed(sub_keys))
    after_ip_d = permute(ct_bits, IP_TABLE)
    L, R = after_ip_d[:32], after_ip_d[32:]
    dec_states = [(list(L), list(R))]
    for i in range(32):
        F     = f_function(R, rev_keys[i], sboxes)
        new_R = xor_bits(L, F)
        L, R  = R, new_R
        dec_states.append((list(L), list(R)))

    pre_fp_d = dec_states[32][1] + dec_states[32][0]
    rv_bits  = permute(pre_fp_d, FP_TABLE)

    return enc_states, dec_states, sub_keys, sboxes, ct_bits, rv_bits


# ============================================================
# [5] IO SHEET  — bit-per-cell, 8×8 IP/FP grids
# ============================================================

def write_io_sheet(ws, pt_hex, key_hex, pt_bits, key_bits,
                   enc_states, dec_states, sub_keys,
                   ct_bits, rv_bits):

    BC = 3        # bit start column
    NBITS = 64
    HEX_COL = BC + NBITS      # hex result column

    _w(ws, 1, 18); _w(ws, 2, 6)
    _narrow_cols(ws, BC, BC + NBITS, 2.2)
    _w(ws, HEX_COL, 12)

    row = 1
    _hdr(ws, row, 1, "IDES — I/O SHEET  (Encryption + Decryption Overview)", fill=_DK_BLUE, span=HEX_COL + 2)
    row += 2

    # ── Inputs ────────────────────────────────────────────────
    _sec(ws, row, 1, "INPUTS", span=HEX_COL + 2); row += 1
    _c(ws, row, 1, "Label", _GREY, _BK_BOLD, _CENTER)
    _c(ws, row, 2, "Hex", _GREY, _BK_BOLD, _CENTER)
    _bit_label_row(ws, row, BC, NBITS)
    _c(ws, row, HEX_COL, "Hex (check)", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, "Plaintext", _LT_BLUE, _BK_BOLD, _LEFT)
    _c(ws, row, 2, pt_hex, _LT_BLUE, _SMALL, _CENTER)
    _bits_row(ws, row, BC, pt_bits, _LT_BLUE)
    _c(ws, row, HEX_COL, _bhex(pt_bits), _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "Key (128-bit)", _CREAM, _BK_BOLD, _LEFT)
    _c(ws, row, 2, key_hex[:16] + "…", _CREAM, _SMALL, _CENTER)
    _bits_row(ws, row, BC, key_bits[:64], _CREAM)
    _c(ws, row, HEX_COL, key_hex[:16], _CREAM, _SMALL, _CENTER)
    row += 1
    _c(ws, row, 1, "Key (bits 65-128)", _CREAM, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, key_bits[64:], _CREAM)
    _c(ws, row, HEX_COL, key_hex[16:], _CREAM, _SMALL, _CENTER)
    row += 2

    # ── IP reference grid ─────────────────────────────────────
    row = _write_ip_grid(ws, row, 1, IP_TABLE, "Initial Permutation (IP) Table — Reference Grid (values = source bit position)")

    # ── IP applied ────────────────────────────────────────────
    _sec(ws, row, 1, "INITIAL PERMUTATION Applied to Plaintext", span=HEX_COL + 2); row += 1
    after_ip = permute(pt_bits, IP_TABLE)
    _c(ws, row, 1, "Label", _GREY, _BK_BOLD, _CENTER)
    _bit_label_row(ws, row, BC, 64)
    _c(ws, row, HEX_COL, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, "Plaintext", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, pt_bits, _LT_BLUE)
    _c(ws, row, HEX_COL, pt_hex, _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "IP(Plaintext)", _YELLOW, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, after_ip, _YELLOW)
    _c(ws, row, HEX_COL, _bhex(after_ip), _YELLOW, _BK_BOLD, _CENTER)
    row += 1

    L0, R0 = after_ip[:32], after_ip[32:]
    _c(ws, row, 1, "L0  (bits 1-32)", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, L0, _GREEN)
    _c(ws, row, BC + 32, "← L0 ends", _GREY, _TINY, _LEFT)
    _c(ws, row, HEX_COL, _bhex(L0), _GREEN, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "R0  (bits 33-64)", _ORANGE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC + 32, R0, _ORANGE)
    _c(ws, row, HEX_COL, _bhex(R0), _ORANGE, _BK_BOLD, _CENTER)
    row += 2

    # ── Encryption round summary ───────────────────────────────
    _sec(ws, row, 1, "ENCRYPTION — 32 Feistel Round States  (Ln | Rn after each round)", span=HEX_COL + 4); row += 1
    for ci, lbl in enumerate(["Rnd", "Ln hex", "Rn hex"], 1):
        _col_hdr(ws, row, ci, lbl)
    _bit_label_row(ws, row, BC, 32)
    _c(ws, row, BC + 32, "‖", _GREY, _TINY, _CENTER)
    _bit_label_row(ws, row, BC + 33, 32)
    row += 1

    for i in range(33):
        L, R = enc_states[i]
        f = _YELLOW if i == 0 else (_GREEN if i == 32 else None)
        _c(ws, row, 1, i,          f, _BK_BOLD if f else _SMALL, _CENTER)
        _c(ws, row, 2, _bhex(L),   f, _SMALL, _CENTER)
        _c(ws, row, 3, _bhex(R),   f, _SMALL, _CENTER)
        _bits_row(ws, row, BC,      L, f)
        _c(ws, row, BC + 32, "‖",  _GREY, _TINY, _CENTER)
        _bits_row(ws, row, BC + 33, R, f)
        row += 1
    row += 1

    # ── Swap + FP ─────────────────────────────────────────────
    _sec(ws, row, 1, "32-bit SWAP then FINAL PERMUTATION  (FP = IP⁻¹)", span=HEX_COL + 2); row += 1
    pre_fp = enc_states[32][1] + enc_states[32][0]
    _c(ws, row, 1, "R32 ‖ L32 (swap)", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, pre_fp, _LT_BLUE)
    _c(ws, row, HEX_COL, _bhex(pre_fp), _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "Ciphertext (FP out)", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, ct_bits, _GREEN)
    _c(ws, row, HEX_COL, _bhex(ct_bits), _GREEN, _BK_BOLD, _CENTER)
    row += 2

    # ── FP reference grid ─────────────────────────────────────
    row = _write_ip_grid(ws, row, 1, FP_TABLE, "Final Permutation (FP = IP⁻¹) Table — Reference Grid")

    # ── Decryption summary ────────────────────────────────────
    _hdr(ws, row, 1, "DECRYPTION — 32 Rounds  (SK32 → SK1)  Ciphertext → Plaintext", fill=_DK_BLUE, span=HEX_COL + 4)
    row += 1
    _c(ws, row, 1, "CT input", _RED_LT, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, ct_bits, _RED_LT)
    _c(ws, row, HEX_COL, _bhex(ct_bits), _RED_LT, _BK_BOLD, _CENTER)
    row += 1

    after_ip_d = permute(ct_bits, IP_TABLE)
    _c(ws, row, 1, "IP(Ciphertext)", _YELLOW, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, after_ip_d, _YELLOW)
    _c(ws, row, HEX_COL, _bhex(after_ip_d), _YELLOW, _BK_BOLD, _CENTER)
    row += 1

    for ci, lbl in enumerate(["Rnd", "Ln hex", "Rn hex"], 1):
        _col_hdr(ws, row, ci, lbl)
    _bit_label_row(ws, row, BC, 32)
    _c(ws, row, BC + 32, "‖", _GREY, _TINY, _CENTER)
    _bit_label_row(ws, row, BC + 33, 32)
    row += 1

    for i in range(33):
        L, R = dec_states[i]
        f = _YELLOW if i == 0 else (_GREEN if i == 32 else None)
        _c(ws, row, 1, i,          f, _BK_BOLD if f else _SMALL, _CENTER)
        _c(ws, row, 2, _bhex(L),   f, _SMALL, _CENTER)
        _c(ws, row, 3, _bhex(R),   f, _SMALL, _CENTER)
        _bits_row(ws, row, BC,      L, f)
        _c(ws, row, BC + 32, "‖",  _GREY, _TINY, _CENTER)
        _bits_row(ws, row, BC + 33, R, f)
        row += 1

    pre_fp_d = dec_states[32][1] + dec_states[32][0]
    _c(ws, row, 1, "R32 ‖ L32 (swap)", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, pre_fp_d, _LT_BLUE)
    _c(ws, row, HEX_COL, _bhex(pre_fp_d), _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "Recovered PT (FP out)", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, rv_bits, _GREEN)
    _c(ws, row, HEX_COL, _bhex(rv_bits), _GREEN, _BK_BOLD, _CENTER)
    row += 1

    match = "✓ MATCH" if rv_bits == pt_bits else "✗ MISMATCH"
    mf = _GREEN if rv_bits == pt_bits else _RED_LT
    _c(ws, row, 1, "Verification", mf, _BK_BOLD, _LEFT)
    ws.merge_cells(start_row=row, start_column=BC, end_row=row, end_column=HEX_COL)
    _c(ws, row, BC, match, mf, Font(bold=True, size=11), _CENTER)

    ws.freeze_panes = "A3"


# ============================================================
# [1][2] KEY SCHEDULE SHEET
# ============================================================

def write_key_schedule_sheet(ws, key_bits, sub_keys):
    BC = 3         # bit start column

    _w(ws, 1, 20); _w(ws, 2, 8)
    _narrow_cols(ws, BC, BC + 127, 2.2)
    _w(ws, BC + 128, 12)
    _w(ws, BC + 129, 14)

    row = 1
    _hdr(ws, row, 1, "IDES KEY SCHEDULE — 32 Sub-key Derivation  (Full Step-by-Step)", fill=_DK_BLUE, span=BC + 50)
    row += 2

    KL, KR = key_bits[:64], key_bits[64:]

    # ══════════════════════════════════════════════════════════
    # STEP 1 — 128-bit master key
    # ══════════════════════════════════════════════════════════
    _sec(ws, row, 1, "Step 1 — 128-bit Master Key  (KL = bits 1-64  |  KR = bits 65-128)", span=BC + 130); row += 1
    _c(ws, row, 1, "Label", _GREY, _BK_BOLD, _CENTER)
    _bit_label_row(ws, row, BC, 128)
    _c(ws, row, BC + 128, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, "Full 128-bit Key", _CREAM, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, key_bits, _CREAM)
    _c(ws, row, BC + 128, _bhex(key_bits), _CREAM, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "KL  (bits 1-64)", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, KL, _LT_BLUE)
    _c(ws, row, BC + 128, _bhex(KL), _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "KR  (bits 65-128)", _ORANGE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC + 64, KR, _ORANGE)
    _c(ws, row, BC + 128, _bhex(KR), _ORANGE, _BK_BOLD, _CENTER)
    row += 2

    # ══════════════════════════════════════════════════════════
    # STEP 2 — PC-1 table grid + application
    # ══════════════════════════════════════════════════════════
    _sec(ws, row, 1, "Step 2 — PC-1 Permutation Table  (selects 56 of 64 bits from each half; drops parity bits at positions 8,16,24,32,40,48,56,64)", span=BC + 130); row += 1

    # PC-1 reference grid (8 rows × 7 cols)
    _c(ws, row, 1, "PC-1  8×7 Grid", _GREY, _BK_BOLD, _CENTER)
    for ci in range(7):
        _col_hdr(ws, row, BC + ci, f"Col {ci+1}")
    row += 1
    for ri in range(8):
        _c(ws, row, 1, f"Row {ri+1}", _GREY, _BK_BOLD, _CENTER)
        for ci in range(7):
            idx = ri * 7 + ci
            v = PC1[idx] if idx < len(PC1) else ""
            _c(ws, row, BC + ci, v, _LT_BLUE, _SMALL, _CENTER, border=True)
        row += 1
    row += 1

    # Apply PC-1 to each half
    _sec(ws, row, 1, "Step 2 — PC-1 Applied  (output: 56 bits per half)", span=BC + 130); row += 1
    _c(ws, row, 1, "Label", _GREY, _BK_BOLD, _CENTER)
    _bit_label_row(ws, row, BC, 56)
    _c(ws, row, BC + 56, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    KL_pc1 = permute(KL, PC1)
    KR_pc1 = permute(KR, PC1)

    _c(ws, row, 1, "PC-1(KL)  56 bits", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, KL_pc1, _LT_BLUE)
    _c(ws, row, BC + 56, _bhex(KL_pc1), _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "PC-1(KR)  56 bits", _ORANGE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, KR_pc1, _ORANGE)
    _c(ws, row, BC + 56, _bhex(KR_pc1), _ORANGE, _BK_BOLD, _CENTER)
    row += 2

    # ══════════════════════════════════════════════════════════
    # [1] 112-bit CONCATENATION ROW — the missing link
    # ══════════════════════════════════════════════════════════
    _sec(ws, row, 1, "Step 2b — Concatenate PC-1(KL) ‖ PC-1(KR) = 112-bit SHA-256 Input  ← this is the seed fed into SHA-256", span=BC + 130, fill=_DK_BLUE); row += 1
    _c(ws, row, 1, "Label", _GREY, _BK_BOLD, _CENTER)
    _bit_label_row(ws, row, BC, 112)
    _c(ws, row, BC + 112, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    concat_112 = KL_pc1 + KR_pc1
    _c(ws, row, 1, "PC-1(KL) ‖ PC-1(KR)", _YELLOW, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, KL_pc1, _LT_BLUE)       # first 56 = KL portion
    _bits_row(ws, row, BC + 56, KR_pc1, _ORANGE)   # next 56  = KR portion
    _c(ws, row, BC + 112, _bhex(concat_112), _YELLOW, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "← KL portion (56b)", _LT_BLUE,  _TINY, _LEFT)
    _c(ws, row, 2, "bits 1-56",           _LT_BLUE,  _TINY, _CENTER)
    row += 1
    _c(ws, row, 1, "← KR portion (56b)", _ORANGE,   _TINY, _LEFT)
    _c(ws, row, 2, "bits 57-112",         _ORANGE,   _TINY, _CENTER)
    row += 2

    # ══════════════════════════════════════════════════════════
    # STEP 3 — SHA-256
    # ══════════════════════════════════════════════════════════
    combined    = bits_to_bytes(concat_112)
    inter_bytes = sha256_hash(combined)
    inter_bits  = []
    for byte in inter_bytes:
        for i in range(7, -1, -1):
            inter_bits.append((byte >> i) & 1)

    master_seed_hex = inter_bytes.hex().upper()

    _sec(ws, row, 1, "Step 3 — SHA-256( 112-bit seed ) → 256-bit Intermediate Key Material", span=BC + 130); row += 1
    _c(ws, row, 1, "SHA-256 input", _YELLOW, _BK_BOLD, _LEFT)
    _c(ws, row, 2, "112 bits (14 bytes) = PC-1(KL) ‖ PC-1(KR) shown above", _YELLOW, _SMALL, _LEFT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=BC + 10)
    row += 1

    _c(ws, row, 1, "SHA-256 output", _PURPLE, _BK_BOLD, _LEFT)
    _c(ws, row, 2, "256 bits (32 bytes) — fixed output size regardless of input size", _PURPLE, _SMALL, _LEFT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=BC + 10)
    row += 1

    _c(ws, row, 1, "Output hex", _PURPLE, _BK_BOLD, _LEFT)
    _c(ws, row, 2, inter_bytes.hex().upper(), _PURPLE, _SMALL, _LEFT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=BC + 60)
    row += 2

    _bit_label_row(ws, row, BC, 128)
    row += 1

    _c(ws, row, 1, "Inter bits 1-128", _PURPLE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, inter_bits[:128], _PURPLE)
    _c(ws, row, BC + 128, _bhex(inter_bits[:128]), _PURPLE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "Inter bits 129-256", _PURPLE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, inter_bits[128:], _PURPLE)
    _c(ws, row, BC + 128, _bhex(inter_bits[128:]), _PURPLE, _BK_BOLD, _CENTER)
    row += 2

    C_state_init = inter_bits[:128]
    D_state_init = inter_bits[128:]

    _sec(ws, row, 1, "Step 3b — Split into C_state (bits 1-128) and D_state (bits 129-256)", span=BC + 130); row += 1
    _bit_label_row(ws, row, BC, 128); row += 1

    _c(ws, row, 1, "C_state  (128 bits)", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, C_state_init, _LT_BLUE)
    _c(ws, row, BC + 128, _bhex(C_state_init), _LT_BLUE, _BK_BOLD, _CENTER)
    row += 1

    _c(ws, row, 1, "D_state  (128 bits)", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, D_state_init, _GREEN)
    _c(ws, row, BC + 128, _bhex(D_state_init), _GREEN, _BK_BOLD, _CENTER)
    row += 2

    # ══════════════════════════════════════════════════════════
    # PC-2 reference grid (before Step 4)
    # ══════════════════════════════════════════════════════════
    _sec(ws, row, 1, "PC-2 Permutation Table  (selects 48 bits from 56-bit C‖D input → sub-key)", span=BC + 130); row += 1
    _c(ws, row, 1, "PC-2  8×6 Grid", _GREY, _BK_BOLD, _CENTER)
    for ci in range(6):
        _col_hdr(ws, row, BC + ci, f"Col {ci+1}")
    row += 1
    for ri in range(8):
        _c(ws, row, 1, f"Row {ri+1}", _GREY, _BK_BOLD, _CENTER)
        for ci in range(6):
            idx = ri * 6 + ci
            v = PC2[idx] if idx < len(PC2) else ""
            _c(ws, row, BC + ci, v, _ORANGE, _SMALL, _CENTER, border=True)
        row += 1
    row += 2

    # ══════════════════════════════════════════════════════════
    # [2] STEP 4 — 32 Sub-key rounds with per-round substeps
    # ══════════════════════════════════════════════════════════
    _sec(ws, row, 1, "Step 4 — 32 Sub-key Generation  (per round: Left-Rotate → Select C[:28] & D[:28] → PC-2 → 48-bit SK)", span=BC + 130, fill=_DK_BLUE); row += 1

    C_st = list(C_state_init)
    D_st = list(D_state_init)

    for i in range(32):
        sh   = SHIFT_SCHEDULE[i]
        C_st = left_rotate(C_st, sh)
        D_st = left_rotate(D_st, sh)
        C28  = C_st[:28]
        D28  = D_st[:28]
        CD56 = C28 + D28
        SK   = permute(CD56, PC2)
        row_fill = _GREEN if i % 2 == 0 else _LT_BLUE

        # Round banner
        _hdr(ws, row, 1,
             f"Round {i+1}  |  Left-shift = {sh}  |  SK{i+1} = PC-2( C{i+1}[:28] ‖ D{i+1}[:28] )",
             fill=_MED_BLUE, span=BC + 130)
        row += 1

        # Bit labels (128 positions)
        _bit_label_row(ws, row, BC, 128); row += 1

        # C_state after rotation (128 bits)
        _c(ws, row, 1, f"C_state after RotL({sh})", row_fill, _BK_BOLD, _LEFT)
        _c(ws, row, 2, f"128b", row_fill, _TINY, _CENTER)
        _bits_row(ws, row, BC, C_st, row_fill)
        _c(ws, row, BC + 128, _bhex(C_st), row_fill, _SMALL, _CENTER)
        row += 1

        # C[:28] selection highlighted
        _c(ws, row, 1, f"C{i+1}[:28] selected", _YELLOW, _BK_BOLD, _LEFT)
        _c(ws, row, 2, "28b", _YELLOW, _TINY, _CENTER)
        _bits_row(ws, row, BC, C28, _YELLOW)
        _c(ws, row, BC + 28, "← only first 28 used", _GREY, _TINY, _LEFT)
        row += 1

        # D_state after rotation (128 bits)
        _c(ws, row, 1, f"D_state after RotL({sh})", row_fill, _BK_BOLD, _LEFT)
        _c(ws, row, 2, "128b", row_fill, _TINY, _CENTER)
        _bits_row(ws, row, BC, D_st, row_fill)
        _c(ws, row, BC + 128, _bhex(D_st), row_fill, _SMALL, _CENTER)
        row += 1

        # D[:28] selection highlighted
        _c(ws, row, 1, f"D{i+1}[:28] selected", _YELLOW, _BK_BOLD, _LEFT)
        _c(ws, row, 2, "28b", _YELLOW, _TINY, _CENTER)
        _bits_row(ws, row, BC, D28, _YELLOW)
        _c(ws, row, BC + 28, "← only first 28 used", _GREY, _TINY, _LEFT)
        row += 1

        # CD56 combined
        _bit_label_row(ws, row, BC, 56)
        row += 1
        _c(ws, row, 1, f"CD56 = C[:28] ‖ D[:28]", _ORANGE, _BK_BOLD, _LEFT)
        _c(ws, row, 2, "56b", _ORANGE, _TINY, _CENTER)
        _bits_row(ws, row, BC, CD56, _ORANGE)
        row += 1

        # SK after PC-2
        _bit_label_row(ws, row, BC, 48)
        row += 1
        _c(ws, row, 1, f"SK{i+1} = PC-2(CD56)", _PURPLE, _BK_BOLD, _LEFT)
        _c(ws, row, 2, "48b", _PURPLE, _TINY, _CENTER)
        _bits_row(ws, row, BC, SK, _PURPLE)
        _c(ws, row, BC + 48, _bhex(SK), _PURPLE, Font(bold=True, size=9), _CENTER)
        row += 2

    ws.freeze_panes = "A3"


# ============================================================
# [3] S-BOX SHEET — full generation trace
# ============================================================

def write_sbox_sheet(ws, key_bits, key_hex):
    from utils import bits_to_bytes, sha256_hash
    from sbox_generator import fisher_yates_shuffle, _row_nonlinearity, NL_THRESHOLD

    _w(ws, 1, 20); _w(ws, 2, 16); _w(ws, 3, 14); _w(ws, 4, 14)
    for c in range(5, 25):
        _w(ws, c, 4.5)
    _w(ws, 25, 8)

    row = 1
    _hdr(ws, row, 1, f"IDES DYNAMIC S-BOXES — Generated from Key: {key_hex}", fill=_DK_BLUE, span=30); row += 1
    _c(ws, row, 1,
       "Each S-box is a 4×16 permutation (values 0-15 once per row). "
       "Generation: SHA-256 seeding → Fisher-Yates shuffle → Walsh-Hadamard NL check (NL ≥ 2).",
       fill=_LT_BLUE, font=Font(italic=True, size=9), align=_LEFT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=30)
    row += 2

    # Master seed derivation
    key_bytes   = bits_to_bytes(key_bits)
    master_seed = sha256_hash(key_bytes)

    _sec(ws, row, 1, "Master Seed Derivation  —  SHA-256( 128-bit key bytes )", span=30); row += 1
    _c(ws, row, 1, "Key (hex)", _LT_BLUE, _BK_BOLD, _LEFT)
    _c(ws, row, 2, key_hex, _LT_BLUE, _SMALL, _LEFT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=20)
    row += 1
    _c(ws, row, 1, "Master Seed (hex)", _PURPLE, _BK_BOLD, _LEFT)
    _c(ws, row, 2, master_seed.hex().upper(), _PURPLE, _SMALL, _LEFT)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=30)
    row += 2

    col_fills = [_GREEN, _LT_BLUE, _ORANGE, _PURPLE, _YELLOW, _RED_LT, _TEAL, _CREAM]

    for s in range(8):
        box_fill = col_fills[s]
        _hdr(ws, row, 1, f"S-Box {s+1}  —  Generation Trace  +  Final 4×16 Table", fill=_DK_BLUE, span=30); row += 1

        final_rows = []

        # ── Generation trace for each of 4 rows ──
        _sec(ws, row, 1, f"S-Box {s+1} — Row Generation Trace", span=30, fill=_MED_BLUE); row += 1
        hdr_cells = ["Row", "Attempt", "Seed Input (hex)", "Row Hash (hex)", "Shuffle Output →", 0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15, "NL Score", "Pass?"]
        for ci, h in enumerate(hdr_cells, 1):
            _col_hdr(ws, row, ci, h)
        row += 1

        for r in range(4):
            attempt = 0
            while True:
                row_seed_input = master_seed + bytes([s, r, attempt])
                row_hash       = sha256_hash(row_seed_input)
                candidate      = fisher_yates_shuffle(row_hash)
                nl             = _row_nonlinearity(candidate)
                passed         = nl >= NL_THRESHOLD
                if passed:
                    final_rows.append(candidate)
                    break
                attempt += 1

            f = box_fill
            _c(ws, row, 1, f"Row {r}", f, _BK_BOLD, _CENTER)
            _c(ws, row, 2, attempt, f, _SMALL, _CENTER)
            _c(ws, row, 3, row_seed_input.hex().upper()[:32] + "…", f, Font(size=7), _LEFT)
            _c(ws, row, 4, row_hash.hex().upper()[:32] + "…", f, Font(size=7), _LEFT)
            _c(ws, row, 5, "→", _GREY, _TINY, _CENTER)
            for ci, v in enumerate(candidate):
                _c(ws, row, 6 + ci, v, f, _SMALL, _CENTER, border=True)
            _c(ws, row, 22, nl, _GREEN if passed else _RED_LT, _BK_BOLD, _CENTER)
            _c(ws, row, 23, "✓ PASS" if passed else "✗ FAIL",
               _GREEN if passed else _RED_LT, Font(bold=True, size=9), _CENTER)
            row += 1
        row += 1

        # ── Final 4×16 S-box grid ──
        _sec(ws, row, 1, f"S-Box {s+1} — Final 4×16 Lookup Table", span=30, fill=_MED_BLUE); row += 1
        _col_hdr(ws, row, 1, "S-Box\\ Col")
        for ci in range(16):
            _col_hdr(ws, row, ci + 2, ci)
        row += 1
        for r in range(4):
            _c(ws, row, 1, f"Row {r}", box_fill, _BK_BOLD, _CENTER)
            for ci in range(16):
                _c(ws, row, ci + 2, final_rows[r][ci], box_fill, _SMALL, _CENTER, border=True)
            row += 1
        row += 2

    ws.freeze_panes = "B3"


# ============================================================
# [4][5] ROUND SHEET — reference-format E expansion + full detail
# ============================================================

def write_round_sheet(ws, round_num, L_prev, R_prev, SK, sboxes, mode="ENC"):
    BC = 3        # bit start column for most sections
    BH = BC + 50  # hex column (beyond 48-bit sections)

    _w(ws, 1, 22); _w(ws, 2, 8)
    _narrow_cols(ws, BC, BC + 50, 2.2)
    _w(ws, BH, 12); _w(ws, BH + 1, 16)

    mode_lbl = "ENCRYPTION" if mode == "ENC" else "DECRYPTION"
    sk_num   = round_num if mode == "ENC" else (33 - round_num)
    sk_label = f"K{round_num}" if mode == "ENC" else f"K{sk_num} (SK{32-(round_num-1)})"

    row = 1
    _hdr(ws, row, 1,
         f"ROUND {round_num}  —  {mode_lbl}  |  Ln = Rn-1   Rn = Ln-1 ⊕ f(Rn-1, Kn)",
         fill=_DK_BLUE, span=BH + 5)
    row += 1
    _c(ws, row, 1, f"n = {round_num}", _GREY, _BK_BOLD, _LEFT)
    _c(ws, row, 2, mode_lbl, _GREY, _SMALL, _LEFT)
    row += 2

    # ── Sub-key ───────────────────────────────────────────────
    _sec(ws, row, 1, f"Sub-key  {sk_label}  (48 bits)", span=BH + 5); row += 1
    _c(ws, row, 1, "Bit #", _GREY, _TINY, _CENTER)
    _bit_label_row(ws, row, BC, 48)
    _c(ws, row, BH, "Hex", _GREY, _TINY, _CENTER)
    row += 1
    _c(ws, row, 1, sk_label, _YELLOW, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, SK, _YELLOW)
    _c(ws, row, BH, _bhex(SK), _YELLOW, Font(bold=True, size=9), _CENTER)
    row += 2

    # ── Rn-1 input ────────────────────────────────────────────
    _sec(ws, row, 1, f"Input  R{round_num-1}  (32 bits)  — goes into E Expansion", span=BH + 5); row += 1
    _c(ws, row, 1, "Bit #", _GREY, _TINY, _CENTER)
    _bit_label_row(ws, row, BC, 32)
    _c(ws, row, BC + 32, "Hex", _GREY, _TINY, _CENTER)
    row += 1
    _c(ws, row, 1, f"R{round_num-1}", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, R_prev, _LT_BLUE)
    _c(ws, row, BC + 32, _bhex(R_prev), _LT_BLUE, Font(bold=True, size=9), _CENTER)
    row += 2

    # ── [4][5] E Expansion — reference-format annotated table ─
    E = permute(R_prev, E_TABLE)
    _sec(ws, row, 1, "E(Rn-1) Expansion  32 → 48 bits  (8 groups of 6 bits each)", span=BH + 5); row += 1

    # Reference-format table headers
    for ci, h in enumerate(["Block", "Borrowed Left\nBit (Col 1)", "Original 4-bit Block\n(Cols 2,3,4,5)", "Borrowed Right\nBit (Col 6)", "R(n-1) with\nBorrowed Bits\n(6-bit result)", "Bit positions\nfrom R_input"], 1):
        _col_hdr(ws, row, ci, h)
    row += 1

    for gi in range(8):
        grp = E[gi * 6:(gi + 1) * 6]
        # Determine source bit positions in R_prev (1-indexed, wrapping)
        src_positions = [E_TABLE[gi * 6 + k] for k in range(6)]
        borrow_left_pos  = src_positions[0]
        orig_positions   = src_positions[1:5]
        borrow_right_pos = src_positions[5]

        f = _GREEN if gi % 2 == 0 else _LT_BLUE
        _c(ws, row, 1, f"Row {gi+1}", f, _BK_BOLD, _CENTER)

        # Borrowed left bit with source annotation
        _c(ws, row, 2,
           f"{grp[0]}  (from bit {borrow_left_pos})",
           f, _SMALL, _CENTER)

        # Original 4-bit middle block
        orig_str = ''.join(str(b) for b in grp[1:5])
        pos_str  = ','.join(str(p) for p in orig_positions)
        _c(ws, row, 3, f"{orig_str}  (bits {pos_str})", f, _SMALL, _CENTER)

        # Borrowed right bit with source annotation
        _c(ws, row, 4,
           f"{grp[5]}  (from bit {borrow_right_pos})",
           f, _SMALL, _CENTER)

        # Full 6-bit result
        _c(ws, row, 5, _bstr(grp), f, Font(bold=True, size=9), _CENTER)

        # Bit-per-cell result
        _c(ws, row, 6, f"b{gi*6+1}-b{gi*6+6}", _GREY, _TINY, _CENTER)
        row += 1

    row += 1
    # Full 48-bit E output bit-per-cell
    _sec(ws, row, 1, "E(Rn-1) Full 48-bit Output", span=BH + 5); row += 1
    _bit_label_row(ws, row, BC, 48); row += 1
    _c(ws, row, 1, "E(Rn-1) full", _ORANGE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, E, _ORANGE)
    _c(ws, row, BH, _bhex(E), _ORANGE, Font(bold=True, size=9), _CENTER)
    row += 2

    # ── XOR with sub-key ──────────────────────────────────────
    XOR48 = xor_bits(SK, E)
    _sec(ws, row, 1, f"Kn ⊕ E(Rn-1)  XOR  (48 bits)", span=BH + 5); row += 1
    _bit_label_row(ws, row, BC, 48)
    _c(ws, row, BH, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, sk_label,      _YELLOW,  _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, SK,    _YELLOW)
    _c(ws, row, BH, _bhex(SK),    _YELLOW,  Font(bold=True, size=9), _CENTER)
    row += 1

    _c(ws, row, 1, "E(Rn-1)",     _ORANGE,  _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, E,     _ORANGE)
    _c(ws, row, BH, _bhex(E),     _ORANGE,  Font(bold=True, size=9), _CENTER)
    row += 1

    _c(ws, row, 1, "XOR result",  _RED_LT,  _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, XOR48, _RED_LT)
    _c(ws, row, BH, _bhex(XOR48), _RED_LT,  Font(bold=True, size=9), _CENTER)
    row += 2

    # ── S-box substitution — full per-group detail ────────────
    _sec(ws, row, 1, "S-Box Substitution  (8 groups of 6 bits → 4-bit output each)", span=BH + 5); row += 1
    for ci, h in enumerate(["S-Box", "b1","b2","b3","b4","b5","b6",
                             "Row bits\n(b1,b6)","Col bits\n(b2-b5)",
                             "Row\nindex","Col\nindex","S-Box\nvalue",
                             "Out bits\n(4-bit)", "o1","o2","o3","o4"], 1):
        _col_hdr(ws, row, ci, h)
    row += 1

    sbox_out_bits = []
    for s in range(8):
        grp     = XOR48[s * 6:(s + 1) * 6]
        row_idx = (grp[0] << 1) | grp[5]
        col_idx = (grp[1] << 3) | (grp[2] << 2) | (grp[3] << 1) | grp[4]
        val     = sboxes[s][row_idx][col_idx]
        out4    = [(val >> 3) & 1, (val >> 2) & 1, (val >> 1) & 1, val & 1]
        sbox_out_bits.extend(out4)

        f = _GREEN if s % 2 == 0 else _LT_BLUE
        _c(ws, row, 1,  f"S{s+1}", f, _BK_BOLD, _CENTER)
        for k, b in enumerate(grp, 2):
            _c(ws, row, k, b, f, _SMALL, _CENTER, border=True)
        _c(ws, row,  8, f"{grp[0]}{grp[5]}",                 f, _SMALL, _CENTER)
        _c(ws, row,  9, f"{grp[1]}{grp[2]}{grp[3]}{grp[4]}", f, _SMALL, _CENTER)
        _c(ws, row, 10, row_idx, f, _SMALL, _CENTER)
        _c(ws, row, 11, col_idx, f, _SMALL, _CENTER)
        _c(ws, row, 12, val,     f, Font(bold=True, size=9), _CENTER)
        _c(ws, row, 13, _bstr(out4), f, Font(bold=True, size=9), _CENTER)
        for k, b in enumerate(out4, 14):
            _c(ws, row, k, b, f, _SMALL, _CENTER, border=True)
        row += 1

    _c(ws, row, 1, "S-Box output (32b)", _GREEN, _BK_BOLD, _LEFT)
    _bit_label_row(ws, row, BC, 32)
    row += 1
    _c(ws, row, 1, "SBox output bits", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, sbox_out_bits, _GREEN)
    _c(ws, row, BC + 32, _bhex(sbox_out_bits), _GREEN, Font(bold=True, size=9), _CENTER)
    row += 2

    # ── P-permutation ─────────────────────────────────────────
    f_bits = permute(sbox_out_bits, P_TABLE)
    _sec(ws, row, 1, "P-Permutation  (32 → 32 bits)", span=BH + 5); row += 1
    _bit_label_row(ws, row, BC, 32)
    _c(ws, row, BC + 32, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, "S-Box output",        _GREEN,  _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, sbox_out_bits, _GREEN)
    _c(ws, row, BC + 32, _bhex(sbox_out_bits), _GREEN, Font(bold=True, size=9), _CENTER)
    row += 1

    _c(ws, row, 1, f"f(R{round_num-1},K{round_num})",  _PURPLE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, f_bits,       _PURPLE)
    _c(ws, row, BC + 32, _bhex(f_bits),  _PURPLE, Font(bold=True, size=9), _CENTER)
    row += 2

    # ── Compute Rn ────────────────────────────────────────────
    R_new = xor_bits(L_prev, f_bits)
    L_new = list(R_prev)

    _sec(ws, row, 1, f"Compute  R{round_num} = L{round_num-1} ⊕ f(R{round_num-1}, K{round_num})", span=BH + 5); row += 1
    _bit_label_row(ws, row, BC, 32)
    _c(ws, row, BC + 32, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, f"L{round_num-1}", _LT_BLUE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, L_prev, _LT_BLUE)
    _c(ws, row, BC + 32, _bhex(L_prev), _LT_BLUE, Font(bold=True, size=9), _CENTER)
    row += 1

    _c(ws, row, 1, f"f(R{round_num-1},K{round_num})", _PURPLE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, f_bits, _PURPLE)
    _c(ws, row, BC + 32, _bhex(f_bits), _PURPLE, Font(bold=True, size=9), _CENTER)
    row += 1

    _c(ws, row, 1, f"R{round_num} = XOR", _ORANGE, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, R_new, _ORANGE)
    _c(ws, row, BC + 32, _bhex(R_new), _ORANGE, Font(bold=True, size=9), _CENTER)
    row += 2

    # ── Round result ──────────────────────────────────────────
    _hdr(ws, row, 1, f"RESULT — Round {round_num} Output", fill=_DK_BLUE, span=BH + 5); row += 1
    _bit_label_row(ws, row, BC, 32)
    _c(ws, row, BC + 32, "Hex", _GREY, _TINY, _CENTER)
    row += 1

    _c(ws, row, 1, f"L{round_num} = R{round_num-1}  (pass-through)", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, L_new, _GREEN)
    _c(ws, row, BC + 32, _bhex(L_new), _GREEN, Font(bold=True, size=9), _CENTER)
    _c(ws, row, BC + 33, "(= R input unchanged)", _GREY, _TINY, _LEFT)
    row += 1

    _c(ws, row, 1, f"R{round_num} = L{round_num-1} ⊕ f  (new R)", _GREEN, _BK_BOLD, _LEFT)
    _bits_row(ws, row, BC, R_new, _GREEN)
    _c(ws, row, BC + 32, _bhex(R_new), _GREEN, Font(bold=True, size=9), _CENTER)
    _c(ws, row, BC + 33, "(XOR result)", _GREY, _TINY, _LEFT)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = True


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 62)
    print("  IDES — Complete Process Visualizer  (Full Reference Format)")
    print("=" * 62)

    pt_hex  = input("  Plaintext  (16 hex chars): ").strip().upper()
    key_hex = input("  Key        (32 hex chars): ").strip().upper()

    if len(pt_hex) != 16:
        print("  ERROR: Plaintext must be exactly 16 hex characters."); return
    if len(key_hex) != 32:
        print("  ERROR: Key must be exactly 32 hex characters."); return

    pt_bits  = hex_to_bits(pt_hex)
    key_bits = hex_to_bits(key_hex)

    print("\n  Computing all round states ...")
    enc_states, dec_states, sub_keys, sboxes, ct_bits, rv_bits = \
        compute_all_states(pt_bits, key_bits)

    ct_hex = _bhex(ct_bits)
    rv_hex = _bhex(rv_bits)
    print(f"  Ciphertext : {ct_hex}")
    print(f"  Recovered  : {rv_hex}")
    print(f"  Match      : {'PASS ✓' if rv_bits == pt_bits else 'FAIL ✗'}")
    print()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  [1] Writing IO SHEET ...")
    ws = wb.create_sheet("IO SHEET")
    write_io_sheet(ws, pt_hex, key_hex, pt_bits, key_bits,
                   enc_states, dec_states, sub_keys, ct_bits, rv_bits)

    print("  [2] Writing KEY SCHEDULE (with 112-bit concat row + per-round substeps) ...")
    ws = wb.create_sheet("KEY SCHEDULE")
    write_key_schedule_sheet(ws, key_bits, sub_keys)

    print("  [3] Writing S-BOXES (with full generation trace) ...")
    ws = wb.create_sheet("S-BOXES")
    write_sbox_sheet(ws, key_bits, key_hex)

    rev_keys = list(reversed(sub_keys))

    for n in range(1, 33):
        L_prev = enc_states[n - 1][0]
        R_prev = enc_states[n - 1][1]
        SK     = sub_keys[n - 1]
        print(f"  [E{n:02d}] Encryption round R{n} ...", end="\r")
        ws = wb.create_sheet(f"R{n}")
        write_round_sheet(ws, n, L_prev, R_prev, SK, sboxes, mode="ENC")
    print("        Encryption rounds R1–R32 done.                    ")

    for n in range(1, 33):
        L_prev = dec_states[n - 1][0]
        R_prev = dec_states[n - 1][1]
        SK     = rev_keys[n - 1]
        print(f"  [D{n:02d}] Decryption round D{n} ...", end="\r")
        ws = wb.create_sheet(f"D{n}")
        write_round_sheet(ws, n, L_prev, R_prev, SK, sboxes, mode="DEC")
    print("        Decryption rounds D1–D32 done.                    ")

    out_file = f"ides_process_{pt_hex}_{key_hex[:8]}.xlsx"
    print(f"\n  Saving → {out_file} ...")
    wb.save(out_file)
    print(f"  Done!  File saved: {out_file}")
    print("=" * 62)


if __name__ == "__main__":
    main()
